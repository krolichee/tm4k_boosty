from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.worksheet.worksheet import Worksheet

from tm4k.wb._names import *
from ._format_config import *
from .formatter import *



def fillEvenOdd(ws: Worksheet):
    fills = (EVEN_ODD_FILLS['even'],EVEN_ODD_FILLS['odd'])
    for row in range(2, ws.max_row):
        d = row%2
        for cell in ws[row]:
            cell.fill = fills[d]


def getColumnDimensionByHeader(ws: Worksheet, header: str) -> ColumnDimension:
    tags_column_letter = int
    first_row = next(ws.rows)
    for cell in first_row:
        if cell.value == header:
            tags_column_letter = cell.column_letter
            break
    else:
        raise
    return ws.column_dimensions[tags_column_letter]


def getColumnNumberByHeader(ws: Worksheet, header: str):
    column_number = 0
    first_row = next(ws.rows)
    for cell in first_row:
        if cell.value == header:
            column_number = cell.column
            break
    else:
        return 0
    return column_number


def getColumnByIndex(ws: Worksheet, index: int):
    return getColumnByNumber(ws, index + 1)


def getColumnByNumber(ws: Worksheet, column_number: int) -> tuple[Cell]:
    return next(ws.iter_cols(min_col=column_number, max_col=column_number, min_row=1, max_row=ws.max_row))


def getColumnByHeader(ws: Worksheet, header: str) -> tuple[Cell]:
    column_number = getColumnNumberByHeader(ws, header)
    return getColumnByNumber(ws, column_number)


def formatNotNullCells(col: tuple[Cell], formatter: Formatter):
    for cell in col:
        if cell.value != "":
            formatter.apply(cell)


def highlightRowWhere(ws: Worksheet, column_header: str, value: str | int):
    return highlightRowWhereIn(ws, column_header, [value])


def highlightRowWhereIn(ws: Worksheet, column_header: str, value_list: list):
    column_number = getColumnNumberByHeader(ws, column_header)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        if row[column_number - 1].value in value_list:
            for cell in row:
                cell.fill = HIGHLIGHT_FILL


def getTagListPaternFillsSet(wb: Workbook):
    result = {}
    ws: Worksheet = wb[TAG_LIST_SHEET_NAME]
    tags_col = getColumnByHeader(ws, "Тег")
    for cell in tags_col[1:]:
        if cell.fill.patternType is not None:
            result[cell.value] = cell.fill.__copy__()
    return result
