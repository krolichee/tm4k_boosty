from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook
from openpyxl.cell.cell import Cell

from tm4k.wb.format._format_config import *
from tm4k.wb._names import *
from openpyxl.styles import *

from openpyxl.styles.fills import PatternFill

from .pallete import getFromPalleteCycle

from .utils import *


def getTagColumnCellsFormatter(header_cell: Cell, pattern_set: dict, color_i: int):
    if header_cell.value in pattern_set.keys():
        formatter_fill: PatternFill = pattern_set[header_cell.value]
    else:
        color = getFromPalleteCycle(color_i)
        formatter_fill = PatternFill(fill_type="solid", fgColor=color)
        color_i += 1

    sides = ['left', 'right', 'top', 'bottom']
    side = Side(border_style='thin', color='000000')
    side_set = {key: side for key in sides}

    formatter = Formatter({
        'fill': formatter_fill,
        'border': Border(**side_set),
        'alignment': Alignment(horizontal='center')
    })
    return formatter, color_i


def colorizeTagMatrixSheet(wb: Workbook):
    ws: Worksheet = wb[TAGS_MATRIX_SHEET_NAME]
    pattern_set = getTagListPaternFillsSet(wb)
    div_col_number = getColumnNumberByHeader(ws, TAGS_MATRIX_DIVIDER_SYMBOL)
    color_i = 0
    for j in range(div_col_number, ws.max_column):
        tag_col = getColumnByIndex(ws, j)
        header_cell = tag_col[0]
        formatter, color_i = getTagColumnCellsFormatter(header_cell, pattern_set, color_i)
        formatter.apply(header_cell)
        formatNotNullCells(tag_col, formatter)


def formatTagListWorksheet(wb: Workbook):
    ws = wb[TAG_LIST_SHEET_NAME]
    for column_header, width in WIDTH_LIST[TAG_LIST_SHEET_NAME].items():
        column_dimension = getColumnDimensionByHeader(ws, column_header)
        column_dimension.width = width
    pass


def formatWorkbook(wb: Workbook):
    formatTagMatrixWorksheet(wb)
    formatTagListWorksheet(wb)


def formatTagDividerCell(cell: Cell):
    TAG_DIVIDER_FORMATTER.apply(cell)


def formatLinkCell(cell):
    LINK_CELL_FORMATTER.apply(cell)
    cell.hyperlink = cell.value


def formatTagMatrixWorksheet(wb: Workbook):
    ws = wb[TAGS_MATRIX_SHEET_NAME]
    for column_header, width in WIDTH_LIST[TAGS_MATRIX_SHEET_NAME].items():
        column_dimension = getColumnDimensionByHeader(ws, column_header)
        column_dimension.width = width

    links_column = getColumnByHeader(ws, 'Ссылка')
    for cell in links_column[1:]:
        formatLinkCell(cell)

    tags_column = getColumnByHeader(ws, TAGS_MATRIX_DIVIDER_SYMBOL)
    for cell in tags_column:
        formatTagDividerCell(cell)

    fillEvenOdd(ws)
    colorizeTagMatrixSheet(wb)
    ws.freeze_panes = ws[2][getColumnNumberByHeader(ws, TAGS_MATRIX_DIVIDER_SYMBOL) - 1]


