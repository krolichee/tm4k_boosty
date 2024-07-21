from openpyxl.styles import *

import tm4k.wb._names as _names
from .formatter import Formatter as _Formatter



WIDTH_LIST = {
    _names.TAGS_MATRIX_SHEET_NAME: {
        "ID": 8, "TS": 2.5,
        "Название": 36,
        "Уровень подписки": 20,
        "Ссылка": 20,
        _names.TAGS_MATRIX_DIVIDER_SYMBOL: 3
    },
    _names.TAG_LIST_SHEET_NAME: {
        "Тег": 20,
        "Комментарий": 52
    }
}
EVEN_ODD_FILLS = {
    'even': PatternFill(fgColor='f0f0f0', fill_type='solid'),
    'odd': PatternFill(fgColor='FFFFFF', fill_type='solid')
}
HIGHLIGHT_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

TAG_DIVIDER_FORMATTER = _Formatter({
    'border': Border(left=Side(border_style='thin', color='000000'), right=Side(border_style='thin', color='000000')),
    'alignment': Alignment(horizontal='center')
})

LINK_CELL_FORMATTER = _Formatter({
    'alignment': Alignment(horizontal='right', shrink_to_fit=True, wrap_text=False),
    'style': "Hyperlink"})
