from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.cell.cell import Cell

from _format_config import *
from ._names import *
from openpyxl.styles import *

from openpyxl.styles.fills import PatternFill
from typing import List

from .pallete import getFromPalleteCycle


def trace(func):
    def wrapper(*args, **kwargs):
        print(f'TRACE: calling {func.__name__}() '
              f'with {args}, {kwargs}')

        original_result = func(*args, **kwargs)

        print(f'TRACE: {func.__name__}() '
              f'returned {original_result!r}')

        return original_result

    return wrapper


class Formatter:
    format_list = {}
    style = {}

    # @trace
    def apply(self, cell: Cell):
        for k, v in self.format_list.items():
            cell.__setattr__(k, v)

    def __init__(self, format_list):
        self.format_list = format_list


def fillColor(cell, color):
    cell.fill = PatternFill(fill_type="solid", fgColor=color)


def formatNotNullCells(col: tuple[Cell], formatter: Formatter):
    for cell in col:
        if cell.value != "":
            formatter.apply(cell)


def getTagListPaternFillsSet(wb: Workbook):
    result = {}
    ws: Worksheet = wb[TAG_LIST_SHEET_NAME]
    tags_col = getColumnByHeader(ws, "Тег")
    for cell in tags_col[1:]:
        if cell.fill.patternType is not None:
            result[cell.value] = cell.fill.__copy__()
    return result


def getTagColumnCellsFormatter(header_cell: Cell, pattern_set: dict, color_i: int):
    if header_cell.value in pattern_set.keys():
        formatter_fill: PatternFill = pattern_set[header_cell.value]
    else:
        color = getFromPalleteCycle(color_i)
        formatter_fill = PatternFill(fill_type="solid", fgColor=color)
        color_i += 1

    sides = ['left', 'right', 'top', 'bottom']
    side = Side(border_style='thin', color='000000')
    side_set = {key: side for key in sides}

    formatter = Formatter({
        'fill': formatter_fill,
        'border': Border(**side_set),
        'alignment': Alignment(horizontal='center')
    })
    return formatter, color_i


def colorizeTagMatrixSheet(wb: Workbook):
    ws: Worksheet = wb[TAGS_MATRIX_SHEET_NAME]
    pattern_set = getTagListPaternFillsSet(wb)
    first_row: list[Cell] = next(ws.rows)
    div_col_number = getColumnNumberByHeader(ws, TAGS_MATRIX_DIVIDER_SYMBOL)
    color_i = 0
    for j in range(div_col_number, ws.max_column):
        tag_col = getColumnByIndex(ws, j)
        header_cell = tag_col[0]
        formatter, color_i = getTagColumnCellsFormatter(header_cell, pattern_set, color_i)
        formatter.apply(header_cell)
        formatNotNullCells(tag_col, formatter)


def highlightRowWhere(ws: Worksheet, column_header: str, value: str | int):
    return highlightRowWhereIn(ws, column_header, [value])


def highlightRowWhereIn(ws: Worksheet, column_header: str, value_list: list):
    column_number = getColumnNumberByHeader(ws, column_header)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        if row[column_number - 1].value in value_list:
            for cell in row:
                cell.fill = HIGHLIGHT_FILL


def formatTagListWorksheet(wb: Workbook):
    ws = wb[TAG_LIST_SHEET_NAME]
    pass


def formatWorkbook(wb: Workbook):
    formatTagMatrixWorksheet(wb)
    formatTagListWorksheet(wb)


def formatTagDividerCell(cell: Cell):
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'))
    cell.border = border
    cell.alignment = Alignment(horizontal='center')


def formatLinkCell(cell):
    cell.alignment = Alignment(horizontal='right', shrink_to_fit=True, wrap_text=False)
    cell.hyperlink = cell.value
    cell.style = "Hyperlink"


def fillEvenOdd(ws: Worksheet):
    even_fill = EVEN_ODD_FILL['even']
    odd_fill = EVEN_ODD_FILL['odd']
    for row in range(2, ws.max_row + 1, 2):
        for cell in ws[row]:
            cell.fill = even_fill
        for cell in ws[row + 1]:
            cell.fill = odd_fill


def formatTagMatrixWorksheet(wb: Workbook):
    ws = wb[TAGS_MATRIX_SHEET_NAME]
    for column_header, width in WIDTH_LIST.items():
        column_dimension = getColumnDimensionByHeader(ws, column_header)
        column_dimension.width = width

    links_column = getColumnByHeader(ws, 'Ссылка')
    for cell in links_column[1:]:
        formatLinkCell(cell)

    tags_column = getColumnByHeader(ws, TAGS_MATRIX_DIVIDER_SYMBOL)
    for cell in tags_column:
        formatTagDividerCell(cell)

    fillEvenOdd(ws)
    colorizeTagMatrixSheet(wb)
    ws.freeze_panes = ws[2][getColumnNumberByHeader(ws, TAGS_MATRIX_DIVIDER_SYMBOL)]


def getColumnDimensionByHeader(ws: Worksheet, header: str) -> ColumnDimension:
    tags_column_letter = int
    first_row = next(ws.rows)
    for cell in first_row:
        if cell.value == header:
            tags_column_letter = cell.column_letter
            break
    else:
        raise
    return ws.column_dimensions[tags_column_letter]


def getColumnNumberByHeader(ws: Worksheet, header: str):
    column_number = 0
    first_row = next(ws.rows)
    for cell in first_row:
        if cell.value == header:
            column_number = cell.column
            break
    else:
        return 0
    return column_number


def getColumnByIndex(ws: Worksheet, index: int):
    return getColumnByNumber(ws, index + 1)


def getColumnByNumber(ws: Worksheet, column_number: int) -> tuple[Cell]:
    return next(ws.iter_cols(min_col=column_number, max_col=column_number, min_row=1, max_row=ws.max_row))


def getColumnByHeader(ws: Worksheet, header: str) -> tuple[Cell]:
    column_number = getColumnNumberByHeader(ws, header)
    return getColumnByNumber(ws, column_number)
