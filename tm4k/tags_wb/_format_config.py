from openpyxl.styles import PatternFill

from ._names import TAGS_MATRIX_DIVIDER_SYMBOL

WIDTH_LIST = {"ID": 8, "TS": 2,
              "Название": 36,
              "Уровень подписки": 20,
              "Ссылка": 20,
              TAGS_MATRIX_DIVIDER_SYMBOL: 3
              }
EVEN_ODD_FILL = {
    'even': PatternFill(fgColor='f0f0f0', fill_type='solid'),
    'odd': PatternFill(fgColor='FFFFFF', fill_type='solid')
}
HIGHLIGHT_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
