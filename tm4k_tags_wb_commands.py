import openpyxl
import pandas as pd

from tm4k_post_field import *
from tm4k_links import *

from tm4k_tags_wb_sheets import *
from tm4k_tags_wb_format import *

from tm4k_post_edit_payload import getPostPayloadWithoutTags
from tm4k_sheet_names import *

from tm4k_file_tags import *

from tm4k_file_blog import *

import requests


def createTagListDf(tag_list: list) -> pd.DataFrame:
    # adjusted_tag_list = adjust_list_size(tag_list,length)
    return pd.DataFrame({'Тег': tag_list, 'Комментарий': [None] * len(tag_list)})


def importTagsFileByBlogId(blog_id: str):
    posts_list = openPostsList(blog_id)
    if posts_list is not None:
        importTagsFileFromBlog(posts_list)


def startfileTagsFile(blog_id: str):
    os.startfile(os.path.abspath(getTagsFilePath(blog_id)))


def importTagsFileFromBlog(posts_list: list):
    blog_id = getPostBlogId(posts_list[0])

    if checkTagsFileExists(blog_id):
        if not messagebox.askquestion("??", "Файл тегов уже существует. "
                                            "Перезапись приведёт к потере неопубликованных тегов и замене матрицы тегов. "
                                            "\nПерезаписать?") == 'yes':
            return
    tag_list = getTagListFromBlog(posts_list)
    tag_list_df = createTagListDf(tag_list)
    tag_matrix_df = getTagMatrixDf(posts_list, tag_list)

    path = getTagsFilePath(blog_id)
    try:
        writer = getTagsFileWriter(blog_id, 'w')
    except:
        messagebox.showwarning("!!", "Нет доступа к файлу тегов. Возможно, от открыт в Excel")
        return

    tag_list_df.to_excel(writer, index=False, sheet_name=TAG_LIST_SHEET_NAME)
    tag_matrix_df.to_excel(writer, index=False, sheet_name=TAGS_MATRIX_SHEET_NAME)
    workbook = writer.book
    formatWorkbook(workbook)
    writer._save()
    startfileTagsFile(blog_id)


def fillDividerColumn(df: pd.DataFrame) -> pd.DataFrame:
    df[TMDS] = df[TMDS].fillna(TMDS)
    return df


def updateTagMatrixTagsByBlogId(blog_id: str):
    writer = getTagsFileWriter(blog_id, 'a', 'replace')
    workbook: Workbook = writer.book
    new_tag_matrix_df = getUpdatedTagMatrixDfFromWorkbook(workbook)
    new_tag_matrix_df = fillDividerColumn(new_tag_matrix_df)
    new_tag_matrix_df.to_excel(excel_writer=writer, sheet_name=TAGS_MATRIX_SHEET_NAME, index=False)
    formatWorkbook(workbook)
    writer._save()


def splitDfByHeader(df: pd.DataFrame, header: str, shift: int = 0) -> tuple[pd.DataFrame, pd.DataFrame]:
    right_index = list(df.keys()).index(header)
    return splitDfByColumnIndex(df, right_index, shift)


def getTagMatrixDfFromWorkbook(wb: Workbook):
    return getDfFromWorksheet(wb[TAGS_MATRIX_SHEET_NAME])


def getTagListDfFromWorkbook(wb: Workbook):
    return getDfFromWorksheet(wb[TAG_LIST_SHEET_NAME])


def splitDfByColumnIndex(df: pd.DataFrame, index: int, shift: int = 0) -> tuple[pd.DataFrame, pd.DataFrame]:
    df1 = df.iloc[:, :(index + shift)]
    df2 = df.iloc[:, (index + shift):]
    return df1, df2


def makeDfByHeaderListAndLength(header_list: list, length: int):
    data = {}
    for header in header_list:
        data[header] = [None] * length
    return pd.DataFrame(data)


def addMissingHeaders(df: pd.DataFrame, header_list: list) -> pd.DataFrame:
    for header in header_list:
        if header not in df.columns:
            df[header] = pd.NA
    return df


def sortDfByHeaderList(df: pd.DataFrame, header_list: list) -> pd.DataFrame:
    all_cols = list(df.keys())
    remaining_cols = [col for col in all_cols if col not in header_list]
    final_order = header_list + remaining_cols
    return df[final_order]


def addAndSortByHeaderList(df: pd.DataFrame, header_list: list):
    df = addMissingHeaders(df, header_list)
    return sortDfByHeaderList(df, header_list)


def replaceNotNullCellsToColumnHeader(df: pd.DataFrame):
    for col in df.columns:
        df[col] = df[col].apply(lambda x: col if pd.notnull(x) else x)
    return df


def splitDfByDividerColumn(df: pd.DataFrame, column_name: str):
    df1, df2_div = splitDfByHeader(df, column_name)
    div_df, df2 = splitDfByHeader(df2_div, column_name, +1)
    return df1, df2, div_df


def getUpdatedTagMatrixDfFromWorkbook(wb: Workbook):
    tag_list = getTagListFromWorkbook(wb)
    tag_matrix_df = getTagMatrixDfFromWorkbook(wb)
    df1, raw_matrix_df, div_df = splitDfByDividerColumn(tag_matrix_df, TAGS_MATRIX_DIVIDER_SYMBOL)
    sorted_raw_matrix_df = addAndSortByHeaderList(raw_matrix_df, tag_list)
    repl_sorted_raw_matrix_df = replaceNotNullCellsToColumnHeader(sorted_raw_matrix_df)
    new_tag_matrix_df = pd.concat([df1,div_df, repl_sorted_raw_matrix_df], axis=1)
    return new_tag_matrix_df


def getPostDict(post: dict):
    result = {"Название": post["title"],
              "Уровень подписки": getSubscrLvlName(post),
              "ID": getPostId(post),
              "TS": getPostTs(post),
              "Ссылка": getPostLink(post)
              }
    for tag_obj in post["tags"]:
        tag = tag_obj["title"]
        result[tag] = tag
    return result


def getPostDf(post):
    post_row_dict = getPostDict(post)
    post_df = pd.DataFrame(post_row_dict, index=[0])
    return post_df


def getTagMatrixDf(posts_list: list, tag_list: list):
    df = pd.DataFrame(columns=TAGS_MATRIX_DF_COLS_LIST + [TAGS_MATRIX_DIVIDER_SYMBOL] + tag_list)
    for post in posts_list:
        post_df = getPostDf(post)
        df = pd.concat([df, post_df], ignore_index=True)
    return df


def getDfFromWorksheet(ws: Worksheet):
    data = ws.values
    columns = next(data)
    df = pd.DataFrame(data, columns=columns)
    return df


def getTagListFromWorkbook(wb: Workbook) -> list:
    tags_df = getTagListDfFromWorkbook(wb)
    tag_list = [tag for tag in list(tags_df["Тег"]) if tag is not None]
    return tag_list


def mergeMatrixPostsOnly(blog_id):
    writer = getTagsFileWriter(blog_id, mode='a', if_sheet_exists='replace')
    wb = writer.book
    m_df_a = getDfFromWorksheet(wb["Матрица тегов A"])
    m_df_b = getDfFromWorksheet(wb["Матрица тегов B"])
    print(m_df_b[TAGS_MATRIX_DF_COLS_LIST])
    print(m_df_b.loc[~m_df_b['ID'].isin(m_df_a['ID']), 'ID'])

    unique_values_list_b = m_df_b.loc[~m_df_b['ID'].isin(m_df_a['ID']), 'ID'].tolist()


def merge(blog_id):
    writer = getTagsFileWriter(blog_id, mode='a', if_sheet_exists='replace')
    wb = writer.book
    m_df_a = getDfFromWorksheet(wb["Матрица тегов A"])
    m_df_b = getDfFromWorksheet(wb["Матрица тегов B"])
    unique_values_list_a = m_df_a.loc[~m_df_a['ID'].isin(m_df_b['ID']), 'ID'].tolist()
    unique_values_list_b = m_df_b.loc[~m_df_b['ID'].isin(m_df_a['ID']), 'ID'].tolist()
    merged = m_df_a.merge(m_df_b, how='right', left_on=["ID"], right_on=["ID"])
    # sorted_merged = merged.sort_values(by="TS")

    merged.to_excel(writer, "Матрица тегов Merged", index=False)
    formatWorkbook(wb)
    writer._save()
    print(unique_values_list_a)
    print(unique_values_list_b)


def updateMatrixPostsByBlogId(blog_id: str):
    writer = getTagsFileWriter(blog_id, 'a')
    wb = writer.book
    matrix_df = getTagMatrixDfFromWorkbook(wb)
    posts_list = openPostsList(blog_id)
    matrix_df_id_list = list(matrix_df["ID"])
    new_ids = []
    for post in posts_list:
        post_id = getPostId(post)
        if post_id not in matrix_df_id_list:
            post_df = getPostDf(post)
            matrix_df = pd.concat([matrix_df, post_df], ignore_index=True, axis='rows')
            new_ids.append(post_id)
    # writer = Writer("/boosty/marcykatya/tags1.xlsx", engine='openpyxl', mode='w')
    print(new_ids)
    matrix_df = matrix_df.sort_values(by='TS', ascending=False)
    matrix_df.to_excel(writer, "updateMatrixPostsByBlogId", index=False)
    wb = writer.book
    highlightRowWhereIn(wb["updateMatrixPostsByBlogId"], "ID", new_ids)
    writer._save()
    startfileTagsFile(blog_id)


def restoreTs(blog_id):
    "Матрица тегов Restored"
    posts_list = openPostsList(blog_id)
    post_ids = {'ID': [],
                'TS': []}
    for post in posts_list:
        id = getPostId(post)
        post_ids['ID'].append(id)
        ts = getPostTs(post)
        post_ids['TS'].append(ts)
    writer = getTagsFileWriter(blog_id, mode='a', if_sheet_exists='replace')
    wb = writer.book
    no_ts_ws = wb[TAGS_MATRIX_SHEET_NAME]
    no_ts_df = getDfFromWorksheet(no_ts_ws)
    post_ids_df = pd.DataFrame(post_ids)
    restoredDf = no_ts_df.merge(post_ids_df, how="left", on="ID")

    missing_ts_mask = restoredDf['TS'].isna()
    restoredDf.loc[missing_ts_mask, 'TS'] = restoredDf.loc[missing_ts_mask, 'ID'].apply(lambda x: getPostTs(
        requests.get(
            getPostApiLink(blog_id, x))
        .json()))
    restoredDf = sortDfByHeaderList(restoredDf, TAGS_MATRIX_DF_COLS_LIST)
    restoredDf = restoredDf.sort_values(by='TS', ascending=False)
    restoredDf.to_excel(writer, "Матрица тегов Restored TS", index=False)
    writer._save()


def publish(blog_id: str, token: str):
    path = getTagsFilePath(blog_id)

    writer = pd.ExcelWriter(path, engine='openpyxl', mode='a')
    wb: Workbook = writer.book
    ws = wb[TAGS_MATRIX_SHEET_NAME]
    df = getDfFromWorksheet(ws)

    headers = {'Authorization': f'Bearer {token}'}
    posts_list = openPostsList(blog_id)
    for post in posts_list:
        post_id = getPostId(post)

        row_df = df.loc[df['ID'] == getPostId(post)]

        tag_list = []

        # todo заадекватить
        _, tags_df = splitDfByHeader(row_df, TAGS_MATRIX_DIVIDER_SYMBOL, +1)

        # todo if tag_cell is not None: tag_cell = tag_cell.header

        for tag_cell in tags_df.iloc[0]:
            print(tag_cell)
            if tag_cell is not None:
                tag_list.append(tag_cell)

        payload = getPostPayloadWithoutTags(post)

        payload['tags'] = ",".join(tag_list)

        url = getPostApiLink(blog_id, post_id)

        response = requests.put(url, data=payload, headers=headers)
        print(response)
        print(response.text)


def getTagListFromBlog(posts_list):
    tag_list = []
    for post in posts_list:
        for tag_obj in post["tags"]:
            tag = tag_obj["title"]
            if tag not in tag_list:
                tag_list.append(tag)
    return tag_list
